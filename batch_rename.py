#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
===============================================================================
批量文件重命名工具
===============================================================================
版本: 1.1.5
作者: CHEN
联系方式: 13516111231
创建日期: 2026-03-13
最后更新: 2026-03-23
===============================================================================
功能描述:
    根据Excel映射表批量重命名指定目录下的文件
    支持递归查找子目录中的文件
    支持文件备份
    支持加密Excel文件（通过WPS/Office COM组件直接读取）
    所有路径均使用相对路径
===============================================================================
"""

import os
import sys
import argparse
import configparser
import shutil
import logging
import tempfile
import time
import csv
from pathlib import Path
from datetime import datetime

# 版本信息
__version__ = '1.1.5'
__author__ = 'CHEN'
__phone__ = '13516111231'
__date__ = '2026-03-23'
__description__ = '批量文件重命名工具 - 根据Excel映射表批量重命名文件（支持加密文件，支持递归查找）'

try:
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException
except ImportError:
    load_workbook = None
    InvalidFileException = Exception

class ExcelEncryptedError(Exception):
    """Excel文件加密异常"""
    pass

class ExcelReader:
    """Excel文件读取器，支持普通和加密文件"""
    
    def __init__(self, config):
        self.config = config
        self.debug_mode = config.getboolean('Settings', 'debug_mode', False)
        self.script_dir = self._get_base_dir()
        
    def _get_base_dir(self):
        """获取程序运行的基础目录"""
        if getattr(sys, 'frozen', False):
            # 打包后的exe
            return os.path.dirname(sys.executable)
        else:
            # 普通Python脚本
            return os.path.dirname(os.path.abspath(__file__))
    
    def _get_absolute_path(self, relative_path):
        """将相对路径转换为绝对路径（基于程序运行目录）"""
        if os.path.isabs(relative_path):
            return relative_path
        return os.path.join(self.script_dir, relative_path)
    
    def read_mapping(self, file_path, has_header=True, root_dir='./', 
                    path_col=0, oldname_col=1, newname_col=2):
        """
        读取Excel映射文件，自动检测是否加密并选择合适的方法
        """
        mappings = []
        
        # 转换文件路径为绝对路径
        abs_file_path = self._get_absolute_path(file_path)
        
        # 转换根目录为绝对路径
        if os.path.isabs(root_dir):
            abs_root_dir = root_dir
        else:
            abs_root_dir = os.path.join(self.script_dir, root_dir)
        
        abs_root_dir = os.path.abspath(abs_root_dir)
        
        if self.debug_mode:
            logging.debug(f"程序运行目录: {self.script_dir}")
            logging.debug(f"映射文件绝对路径: {abs_file_path}")
            logging.debug(f"根目录绝对路径: {abs_root_dir}")
        
        # 先尝试用openpyxl读取普通文件
        try:
            if self.debug_mode:
                logging.debug(f"尝试用openpyxl读取文件: {abs_file_path}")
            
            mappings = self._read_with_openpyxl(abs_file_path, has_header, abs_root_dir, 
                                                path_col, oldname_col, newname_col)
            
            # 如果成功读取到数据，直接返回
            if mappings:
                logging.info(f"成功使用openpyxl读取映射表，共 {len(mappings)} 条规则")
                return mappings
                
        except Exception as e:
            error_msg = str(e).lower()
            if self.debug_mode:
                logging.debug(f"openpyxl读取失败: {error_msg}")
            
            # 检测是否为加密文件
            is_encrypted = any(keyword in error_msg for keyword in [
                'encrypted', 'password', '解密', '加密', 
                'zipfile', 'badzipfile', 'not a zip file', 
                'unsupported format', 'corrupt'
            ])
            
            if is_encrypted:
                logging.warning(f"检测到加密Excel文件: {os.path.basename(file_path)}")
                logging.info("将尝试使用WPS/Office COM组件直接读取（自动解密）...")
            else:
                logging.warning(f"openpyxl读取失败: {e}，将尝试使用COM组件...")
        
        # 尝试用COM组件直接读取
        try:
            use_com = self.config.getboolean('Settings', 'use_com_for_encrypted', True)
            if not use_com:
                raise ExcelEncryptedError("文件读取失败且未启用COM组件读取")
            
            com_app = self.config.get('Settings', 'com_application', 'auto').lower()
            
            if self.debug_mode:
                logging.debug(f"尝试使用COM组件直接读取，应用: {com_app}")
            
            mappings = self._read_with_com_direct(abs_file_path, has_header, abs_root_dir,
                                                 path_col, oldname_col, newname_col, com_app)
            
            if mappings:
                logging.info(f"成功使用COM组件直接读取映射表，共 {len(mappings)} 条规则")
                return mappings
            else:
                raise Exception("COM组件读取返回空数据")
                
        except ImportError:
            logging.error("使用COM组件需要安装pywin32")
            logging.info("请运行: pip install pywin32")
            raise ExcelEncryptedError("缺少pywin32模块")
        except Exception as e:
            logging.error(f"COM组件读取失败: {e}")
            raise ExcelEncryptedError(f"所有读取方法均失败: {e}")
    
    def _read_with_openpyxl(self, file_path, has_header, root_dir, path_col, oldname_col, newname_col):
        """使用openpyxl读取普通Excel文件"""
        if load_workbook is None:
            logging.error("处理 .xlsx 文件需要安装 openpyxl。请运行：pip install openpyxl")
            sys.exit(1)
        
        mappings = []
        
        try:
            # 尝试以只读模式打开
            wb = load_workbook(file_path, data_only=True, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            
            if not rows:
                wb.close()
                return mappings
            
            start_row = 1 if has_header else 0
            
            if self.debug_mode:
                logging.debug(f"读取XLSX文件，从第{start_row+1}行开始")
                logging.debug(f"列配置: 路径列={path_col}, 原名列={oldname_col}, 新名列={newname_col}")
            
            for i, row in enumerate(rows[start_row:], start=start_row+1):
                if len(row) > max(path_col, oldname_col, newname_col):
                    dir_path_str = str(row[path_col]).strip() if row[path_col] else ''
                    old_name = str(row[oldname_col]).strip() if row[oldname_col] else ''
                    new_name = str(row[newname_col]).strip() if row[newname_col] else ''
                    
                    if dir_path_str and old_name and new_name:
                        # 处理目录路径
                        if os.path.isabs(dir_path_str):
                            full_dir_path = dir_path_str
                        else:
                            full_dir_path = os.path.join(root_dir, dir_path_str)
                        
                        # 规范化路径
                        full_dir_path = os.path.normpath(full_dir_path)
                        
                        mappings.append((full_dir_path, old_name, new_name))
                        
                        if self.debug_mode:
                            logging.debug(f"第{i}行: 目录={full_dir_path}, 原文件={old_name}, 新文件={new_name}")
            
            wb.close()
            
        except Exception as e:
            if 'encrypted' in str(e).lower() or 'badzipfile' in str(e).lower() or 'not a zip file' in str(e).lower():
                raise ExcelEncryptedError(f"文件可能已加密: {e}")
            else:
                raise
        
        return mappings
    
    def _read_with_com_direct(self, file_path, has_header, root_dir, path_col, oldname_col, newname_col, com_app='auto'):
        """
        使用COM组件直接读取Excel文件内容（不通过CSV中间文件）
        这种方法可以确保加密文件被自动解密
        """
        mappings = []
        
        try:
            # 先检查是否已安装pywin32
            try:
                import win32com.client
                from win32com.client import Dispatch
                import pythoncom
            except ImportError as e:
                logging.error(f"请先安装pywin32: pip install pywin32")
                raise ImportError("缺少pywin32模块，请安装后重试")

            # 初始化COM
            pythoncom.CoInitialize()
            
            try:
                # 根据系统选择COM组件
                excel_app = None
                
                if com_app == 'auto':
                    # 先尝试Excel，再尝试WPS
                    if self.debug_mode:
                        logging.debug("自动模式：先尝试Excel...")
                    excel_app = self._connect_excel()
                    if not excel_app:
                        if self.debug_mode:
                            logging.debug("Excel失败，尝试WPS...")
                        excel_app = self._connect_wps()
                elif com_app == 'excel':
                    excel_app = self._connect_excel()
                elif com_app == 'wps':
                    excel_app = self._connect_wps()
                
                if not excel_app:
                    raise Exception("无法连接到Excel或WPS应用程序")
                
                # 设置应用程序属性
                excel_app.Visible = False
                excel_app.DisplayAlerts = False
                
                try:
                    # 打开工作簿 - 这一步会触发加密软件自动解密
                    if self.debug_mode:
                        logging.debug(f"正在打开文件（触发自动解密）: {file_path}")
                    
                    # 尝试以读写模式打开
                    wb = excel_app.Workbooks.Open(file_path)
                    
                    # 等待一下，确保文件完全加载和解密
                    time.sleep(0.5)
                    
                    # 获取活动工作表
                    ws = wb.ActiveSheet
                    
                    # 获取使用的范围
                    used_range = ws.UsedRange
                    
                    if used_range is None:
                        if self.debug_mode:
                            logging.debug("工作表为空")
                        wb.Close()
                        excel_app.Quit()
                        return mappings
                    
                    # 获取行数和列数
                    rows = used_range.Rows.Count
                    cols = used_range.Columns.Count
                    
                    if self.debug_mode:
                        logging.debug(f"工作表大小: {rows}行 x {cols}列")
                    
                    # 确定起始行
                    start_row = 2 if has_header else 1
                    
                    # 逐行读取数据
                    for row_num in range(start_row, rows + 1):
                        try:
                            # 读取指定列的数据
                            dir_cell = ws.Cells(row_num, path_col + 1)  # COM组件列从1开始
                            old_cell = ws.Cells(row_num, oldname_col + 1)
                            new_cell = ws.Cells(row_num, newname_col + 1)
                            
                            # 获取值
                            dir_path_str = str(dir_cell.Value).strip() if dir_cell.Value else ''
                            old_name = str(old_cell.Value).strip() if old_cell.Value else ''
                            new_name = str(new_cell.Value).strip() if new_cell.Value else ''
                            
                            if dir_path_str and old_name and new_name:
                                # 处理目录路径
                                if os.path.isabs(dir_path_str):
                                    full_dir_path = dir_path_str
                                else:
                                    full_dir_path = os.path.join(root_dir, dir_path_str)
                                
                                # 规范化路径
                                full_dir_path = os.path.normpath(full_dir_path)
                                
                                mappings.append((full_dir_path, old_name, new_name))
                                
                                if self.debug_mode and len(mappings) <= 5:
                                    logging.debug(f"第{row_num}行: 目录={full_dir_path}, 原文件={old_name}, 新文件={new_name}")
                            
                        except Exception as e:
                            if self.debug_mode:
                                logging.debug(f"读取第{row_num}行失败: {e}")
                            continue
                    
                    # 关闭工作簿
                    wb.Close()
                    
                    if self.debug_mode:
                        logging.debug(f"成功读取 {len(mappings)} 条映射规则")
                    
                except Exception as e:
                    logging.error(f"读取Excel内容失败: {e}")
                    raise
                finally:
                    # 退出应用程序
                    try:
                        excel_app.Quit()
                    except:
                        pass
                
            finally:
                # 清理COM
                pythoncom.CoUninitialize()
            
        except Exception as e:
            logging.error(f"COM组件直接读取失败: {e}")
            raise
        
        return mappings
    
    def _connect_excel(self):
        """连接到Excel应用程序"""
        try:
            import win32com.client
            if self.debug_mode:
                logging.debug("尝试连接 Excel.Application...")
            excel = win32com.client.Dispatch("Excel.Application")
            if self.debug_mode:
                logging.debug("成功连接 Excel.Application")
            return excel
        except Exception as e:
            if self.debug_mode:
                logging.debug(f"连接Excel失败: {e}")
            return None
    
    def _connect_wps(self):
        """连接到WPS应用程序"""
        try:
            import win32com.client
            
            # 尝试多种可能的WPS ProgID
            prog_ids = ["Kwps.Application", "Wps.Application", "ET.Application"]
            
            for prog_id in prog_ids:
                try:
                    if self.debug_mode:
                        logging.debug(f"尝试连接 {prog_id}...")
                    wps = win32com.client.Dispatch(prog_id)
                    if self.debug_mode:
                        logging.debug(f"成功连接 {prog_id}")
                    return wps
                except:
                    continue
            
            return None
            
        except Exception as e:
            if self.debug_mode:
                logging.debug(f"连接WPS失败: {e}")
            return None


class Config:
    """配置管理类"""
    def __init__(self, config_file='config.ini'):
        self.base_dir = self._get_base_dir()
        self.config_file = self._get_absolute_path(config_file)
        self.config = configparser.ConfigParser()
        self.load_config()
    
    def _get_base_dir(self):
        """获取程序运行的基础目录"""
        if getattr(sys, 'frozen', False):
            # 打包后的exe
            return os.path.dirname(sys.executable)
        else:
            # 普通Python脚本
            return os.path.dirname(os.path.abspath(__file__))
    
    def _get_absolute_path(self, relative_path):
        """将相对路径转换为绝对路径（基于程序运行目录）"""
        if os.path.isabs(relative_path):
            return relative_path
        return os.path.join(self.base_dir, relative_path)
    
    def _get_relative_path(self, absolute_path):
        """将绝对路径转换为相对于程序运行目录的路径"""
        try:
            return os.path.relpath(absolute_path, self.base_dir)
        except:
            return absolute_path
    
    def load_config(self):
        """加载配置文件"""
        if os.path.exists(self.config_file):
            self.config.read(self.config_file, encoding='utf-8')
        else:
            self.create_default_config()
    
    def create_default_config(self):
        """创建默认配置文件（使用相对路径）"""
        self.config['Paths'] = {
            'mapping_file': 'rename.xlsx',
            'root_dir': '.',
            'backup_dir': 'backup'
        }
        self.config['Settings'] = {
            'has_header': 'true',
            'auto_execute': 'false',
            'sanitize_filename': 'true',
            'illegal_char_replacement': '_',
            'create_backup': 'false',
            'path_column': '0',
            'oldname_column': '1',
            'newname_column': '2',
            'debug_mode': 'true',
            'use_com_for_encrypted': 'true',
            'com_application': 'auto'  # auto, wps, excel
        }
        self.config['Logging'] = {
            'enable_logging': 'true',
            'log_file': 'rename_log.txt',
            'log_level': 'DEBUG'
        }
        
        # 确保配置目录存在
        config_dir = os.path.dirname(self.config_file)
        if config_dir and not os.path.exists(config_dir):
            os.makedirs(config_dir)
        
        with open(self.config_file, 'w', encoding='utf-8') as f:
            self.config.write(f)
        print(f"已创建默认配置文件: {self._get_relative_path(self.config_file)}")
    
    def get(self, section, key, fallback=None):
        """获取配置值"""
        try:
            return self.config.get(section, key, fallback=fallback)
        except:
            return fallback
    
    def getint(self, section, key, fallback=0):
        """获取整型配置值"""
        try:
            return self.config.getint(section, key, fallback=fallback)
        except:
            try:
                return int(self.config.get(section, key, fallback=fallback))
            except:
                return fallback
    
    def getboolean(self, section, key, fallback=False):
        """获取布尔型配置值"""
        try:
            return self.config.getboolean(section, key, fallback=fallback)
        except:
            return fallback
    
    def get_path(self, section, key, fallback='.'):
        """获取路径配置（返回绝对路径）"""
        path = self.get(section, key, fallback)
        return self._get_absolute_path(path)
    
    def get_relative_path(self, section, key, fallback='.'):
        """获取相对路径配置（返回相对于程序运行目录的路径）"""
        path = self.get(section, key, fallback)
        abs_path = self._get_absolute_path(path)
        return self._get_relative_path(abs_path)

def show_version():
    """显示版本信息"""
    version_info = f"""
===============================================================================
                    批量文件重命名工具 v{__version__}
===============================================================================
版本: {__version__}
作者: {__author__}
电话: {__phone__}
日期: {__date__}
描述: {__description__}
===============================================================================
"""
    print(version_info)

def setup_logging(config):
    """设置日志系统"""
    if config.getboolean('Logging', 'enable_logging', fallback=True):
        log_file = config.get_path('Logging', 'log_file', 'rename_log.txt')
        log_level = config.get('Logging', 'log_level', 'INFO')
        
        log_dir = os.path.dirname(log_file)
        if log_dir and not os.path.exists(log_dir):
            os.makedirs(log_dir)
        
        level_map = {
            'DEBUG': logging.DEBUG,
            'INFO': logging.INFO,
            'WARNING': logging.WARNING,
            'ERROR': logging.ERROR
        }
        
        logging.basicConfig(
            level=level_map.get(log_level, logging.INFO),
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_file, encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
        
        # 记录路径信息
        logging.info(f"程序运行目录: {config.base_dir}")
        logging.info(f"日志文件: {os.path.basename(log_file)}")
    else:
        logging.basicConfig(
            level=logging.INFO,
            format='%(message)s',
            handlers=[logging.StreamHandler()]
        )

def find_file_recursive(directory, filename, debug_mode=False):
    """
    递归查找文件
    返回找到的文件完整路径，如果没有找到返回None
    """
    if debug_mode:
        logging.debug(f"在目录 {directory} 中递归查找文件: {filename}")
    
    if not os.path.exists(directory):
        if debug_mode:
            logging.debug(f"目录不存在: {directory}")
        return None
    
    # 使用 os.walk 递归遍历所有子目录
    for root, dirs, files in os.walk(directory):
        if filename in files:
            found_path = os.path.join(root, filename)
            if debug_mode:
                logging.debug(f"在子目录中找到文件: {found_path}")
            return found_path
    
    if debug_mode:
        logging.debug(f"未找到文件: {filename}")
    return None

def get_file_info(directory, old_name, debug_mode=False, recursive_find=False):
    """
    获取文件的完整信息，验证文件是否存在
    支持递归查找子目录中的文件
    返回文件信息字典
    """
    # 构建完整的旧文件路径
    old_full_path = os.path.join(directory, old_name)
    exists = os.path.isfile(old_full_path)
    actual_directory = directory
    actual_path = old_full_path
    
    # 如果当前目录不存在文件，且启用了递归查找
    if not exists and recursive_find:
        found_path = find_file_recursive(directory, old_name, debug_mode)
        if found_path:
            actual_path = found_path
            actual_directory = os.path.dirname(found_path)
            exists = True
            if debug_mode:
                logging.debug(f"通过递归找到文件: {found_path}")
    
    if debug_mode:
        logging.debug(f"检查文件信息:")
        logging.debug(f"  目录: {actual_directory}")
        logging.debug(f"  文件名: {old_name}")
        logging.debug(f"  完整路径: {actual_path}")
        logging.debug(f"  文件存在: {exists}")
        
        if not exists:
            if os.path.exists(actual_directory):
                logging.debug(f"  目录存在，列出前10个文件:")
                try:
                    files = os.listdir(actual_directory)[:10]
                    for f in files:
                        logging.debug(f"    - {f}")
                except Exception as e:
                    logging.debug(f"  读取目录失败: {e}")
            else:
                logging.debug(f"  目录不存在: {actual_directory}")
    
    return {
        'directory': actual_directory,
        'old_name': old_name,
        'old_full_path': actual_path,
        'exists': exists
    }

def sanitize_filename(filename, replacement='_'):
    """
    清理文件名中的非法字符
    """
    illegal_chars = r'\/:*?"<>|'
    for ch in illegal_chars:
        filename = filename.replace(ch, replacement)
    
    filename = filename.strip('. ')
    return filename

def create_backup(file_path, backup_dir):
    """创建文件备份"""
    if not os.path.exists(backup_dir):
        os.makedirs(backup_dir)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = os.path.basename(file_path)
    backup_name = f"{timestamp}_{filename}"
    backup_path = os.path.join(backup_dir, backup_name)
    
    try:
        counter = 1
        while os.path.exists(backup_path):
            name, ext = os.path.splitext(backup_name)
            backup_path = os.path.join(backup_dir, f"{name}_{counter}{ext}")
            counter += 1
        
        shutil.copy2(file_path, backup_path)
        logging.info(f"已创建备份: {backup_path}")
        return True
    except Exception as e:
        logging.error(f"创建备份失败: {e}")
        return False

def check_required_files(base_dir):
    """检查必要的文件是否存在"""
    config_file = os.path.join(base_dir, 'config.ini')
    mapping_file = os.path.join(base_dir, 'rename.xlsx')
    
    missing_files = []
    
    if not os.path.exists(config_file):
        missing_files.append('config.ini')
    
    if not os.path.exists(mapping_file):
        missing_files.append('rename.xlsx')
    
    if missing_files:
        print("\n" + "="*60)
        print("错误：以下必要文件不存在：")
        for file in missing_files:
            print(f"  - {file}")
        print(f"\n当前目录: {base_dir}")
        print("="*60)
        return False
    
    return True

def main():
    # 显示版本信息
    show_version()
    
    # 获取程序运行的基础目录
    if getattr(sys, 'frozen', False):
        # 打包后的exe
        base_dir = os.path.dirname(sys.executable)
    else:
        # 普通Python脚本
        base_dir = os.path.dirname(os.path.abspath(__file__))
    
    # 切换到程序运行目录
    os.chdir(base_dir)
    
    # 检查必要文件
    if not check_required_files(base_dir):
        input("\n按回车键退出...")
        return
    
    parser = argparse.ArgumentParser(description=__description__)
    parser.add_argument('--config', '-c', default='config.ini',
                        help='配置文件路径（默认：config.ini，支持相对路径）')
    parser.add_argument('--mapping', '-m',
                        help='映射表文件路径（优先级高于配置文件，支持相对路径）')
    parser.add_argument('--root-dir', '-r',
                        help='根目录路径（用于解析相对路径，优先级高于配置文件，支持相对路径）')
    parser.add_argument('--no-header', action='store_true',
                        help='映射表没有标题行（优先级高于配置文件）')
    parser.add_argument('--execute', action='store_true',
                        help='直接执行重命名而不询问')
    parser.add_argument('--debug', action='store_true',
                        help='启用调试模式')
    parser.add_argument('--recursive-find', action='store_true',
                        help='递归查找文件（在子目录中查找）')
    parser.add_argument('--version', '-v', action='store_true',
                        help='显示版本信息')
    
    args = parser.parse_args()
    
    # 如果只是显示版本信息，则退出
    if args.version:
        input("\n按回车键退出...")
        return
    
    # 加载配置
    config = Config(args.config)
    
    # 设置日志
    setup_logging(config)
    
    # 获取配置值（转换为绝对路径）
    mapping_file = args.mapping if args.mapping else config.get('Paths', 'mapping_file', 'rename.xlsx')
    
    # 如果不是绝对路径，则基于base_dir构建
    if not os.path.isabs(mapping_file):
        mapping_file = os.path.join(base_dir, mapping_file)
    
    root_dir = args.root_dir if args.root_dir else config.get('Paths', 'root_dir', '.')
    if not os.path.isabs(root_dir):
        root_dir = os.path.join(base_dir, root_dir)
    
    has_header = not args.no_header if args.no_header else config.getboolean('Settings', 'has_header', True)
    auto_execute = args.execute or config.getboolean('Settings', 'auto_execute', False)
    sanitize = config.getboolean('Settings', 'sanitize_filename', True)
    replacement = config.get('Settings', 'illegal_char_replacement', '_')
    create_backup_flag = config.getboolean('Settings', 'create_backup', False)
    backup_dir = config.get_path('Settings', 'backup_dir', 'backup')
    debug_mode = args.debug or config.getboolean('Settings', 'debug_mode', False)
    recursive_find = args.recursive_find or False
    
    # 列索引配置
    path_col = config.getint('Settings', 'path_column', 0)
    oldname_col = config.getint('Settings', 'oldname_column', 1)
    newname_col = config.getint('Settings', 'newname_column', 2)
    
    # 检查映射文件是否存在
    if not os.path.exists(mapping_file):
        logging.error(f"映射表文件不存在: {os.path.basename(mapping_file)}")
        logging.info(f"请确保 {os.path.basename(mapping_file)} 文件与程序在同一目录")
        input("\n按回车键退出...")
        return
    
    # 显示路径信息
    logging.info(f"使用映射表: {os.path.basename(mapping_file)}")
    logging.info(f"根目录: {os.path.relpath(root_dir, base_dir)}")
    logging.info(f"备份目录: {os.path.relpath(backup_dir, base_dir)}")
    
    if debug_mode:
        logging.info("调试模式已启用")
        logging.info(f"程序目录: {base_dir}")
    
    # 创建Excel读取器
    excel_reader = ExcelReader(config)
    
    # 读取映射
    try:
        raw_mappings = excel_reader.read_mapping(
            mapping_file, 
            has_header, 
            root_dir,
            path_col, 
            oldname_col, 
            newname_col
        )
    except ExcelEncryptedError as e:
        logging.error(f"读取加密Excel文件失败: {e}")
        logging.info("\n提示:")
        logging.info("1. 请确保已安装pywin32: pip install pywin32")
        logging.info("2. 请确保系统已安装WPS或Microsoft Office")
        logging.info("3. 检查配置文件中的 com_application 设置")
        logging.info("4. 确保加密软件已正常运行，WPS/Office可以自动解密文件")
        input("\n按回车键退出...")
        return
    except Exception as e:
        logging.error(f"读取映射表失败: {e}")
        input("\n按回车键退出...")
        return
    
    if not raw_mappings:
        logging.warning("映射表为空，请检查文件内容")
        input("\n按回车键退出...")
        return
    
    # 验证每个文件并准备重命名操作
    rename_ops = []      # 可执行的重命名操作
    not_found = []       # 文件不存在的记录
    conflicts = []       # 目标文件已存在的冲突
    invalid_paths = []   # 路径无效的记录
    found_by_recursive = []  # 通过递归找到的文件
    
    for i, (directory, old_name, new_name) in enumerate(raw_mappings, 1):
        if debug_mode:
            logging.debug(f"\n{'='*50}")
            logging.debug(f"处理第{i}条映射:")
            logging.debug(f"  目录: {directory}")
            logging.debug(f"  原文件: {old_name}")
            logging.debug(f"  新文件: {new_name}")
        
        # 检查目录是否存在
        if not os.path.isdir(directory):
            invalid_paths.append((directory, old_name, new_name, "目录不存在"))
            continue
        
        # 获取文件信息（支持递归查找）
        file_info = get_file_info(directory, old_name, debug_mode, recursive_find)
        
        if not file_info['exists']:
            not_found.append((directory, old_name, new_name))
            continue
        
        # 如果通过递归找到文件，记录日志
        if file_info['directory'] != directory:
            found_by_recursive.append((old_name, file_info['old_full_path']))
            if debug_mode:
                logging.debug(f"通过递归找到文件: {file_info['old_full_path']}")
        
        # 清理新文件名
        if sanitize:
            new_name_clean = sanitize_filename(new_name, replacement)
        else:
            new_name_clean = new_name
        
        # 构建新文件路径（使用文件实际所在的目录）
        new_full_path = os.path.join(file_info['directory'], new_name_clean)
        
        # 如果新旧路径相同，跳过
        if file_info['old_full_path'] == new_full_path:
            if debug_mode:
                logging.debug(f"跳过: {old_name} 新文件名相同")
            continue
        
        # 检查目标文件是否已存在
        if os.path.exists(new_full_path):
            conflicts.append((file_info['directory'], old_name, new_name_clean))
            continue
        
        rename_ops.append({
            'directory': file_info['directory'],
            'old_name': old_name,
            'new_name': new_name_clean,
            'old_path': file_info['old_full_path'],
            'new_path': new_full_path,
            'original_new': new_name
        })
        
        if debug_mode:
            logging.debug(f"添加重命名操作: {old_name} -> {new_name_clean}")
    
    # 预览信息
    logging.info("\n" + "="*60)
    logging.info("预览重命名操作：")
    
    if rename_ops:
        ops_by_dir = {}
        for op in rename_ops:
            dir_name = op['directory']
            if dir_name not in ops_by_dir:
                ops_by_dir[dir_name] = []
            ops_by_dir[dir_name].append((op['old_name'], op['new_name'], op['original_new']))
        
        for dir_name, ops in ops_by_dir.items():
            # 显示相对路径
            rel_dir = os.path.relpath(dir_name, base_dir)
            logging.info(f"\n目录: {rel_dir}")
            for old_name, new_name, original_new in ops:
                if original_new != new_name:
                    logging.info(f'  "{old_name}" -> "{new_name}" (原计划: "{original_new}")')
                else:
                    logging.info(f'  "{old_name}" -> "{new_name}"')
    else:
        logging.info("  没有需要重命名的文件")
    
    if found_by_recursive:
        logging.info(f"\n通过递归找到的文件 ({len(found_by_recursive)} 个):")
        for old_name, found_path in found_by_recursive[:10]:
            rel_path = os.path.relpath(found_path, base_dir)
            logging.info(f'  "{old_name}" -> {rel_path}')
        if len(found_by_recursive) > 10:
            logging.info(f'  ... 还有 {len(found_by_recursive) - 10} 个文件')
    
    if not_found:
        logging.warning(f"\n以下 {len(not_found)} 个文件不存在，将跳过：")
        for directory, old_name, new_name in not_found[:10]:
            rel_dir = os.path.relpath(directory, base_dir)
            logging.warning(f'  目录: {rel_dir}')
            logging.warning(f'  文件: "{old_name}" -> "{new_name}"')
        if len(not_found) > 10:
            logging.warning(f'  ... 还有 {len(not_found) - 10} 个文件')
    
    if invalid_paths:
        logging.warning(f"\n以下 {len(invalid_paths)} 个目录无效：")
        for directory, old_name, new_name, reason in invalid_paths[:10]:
            rel_dir = os.path.relpath(directory, base_dir)
            logging.warning(f'  目录: {rel_dir} - {reason}')
        if len(invalid_paths) > 10:
            logging.warning(f'  ... 还有 {len(invalid_paths) - 10} 个目录')
    
    if conflicts:
        logging.warning(f"\n以下 {len(conflicts)} 个新文件名已存在，将跳过：")
        for dir_name, old_name, new_name in conflicts[:10]:
            rel_dir = os.path.relpath(dir_name, base_dir)
            logging.warning(f'  目录: {rel_dir}')
            logging.warning(f'  "{old_name}" -> "{new_name}" (目标已存在)')
        if len(conflicts) > 10:
            logging.warning(f'  ... 还有 {len(conflicts) - 10} 个冲突')
    
    # 执行或退出
    if not rename_ops:
        logging.info("\n没有可执行的重命名操作，退出。")
        
        if debug_mode:
            logging.info("\n调试信息汇总：")
            logging.info(f"总映射数: {len(raw_mappings)}")
            logging.info(f"文件不存在: {len(not_found)}")
            logging.info(f"目录无效: {len(invalid_paths)}")
            logging.info(f"文件名冲突: {len(conflicts)}")
            logging.info(f"递归找到: {len(found_by_recursive)}")
        
        input("\n按回车键退出...")
        return
    

    
    # 强制自动执行，不等待用户输入
    confirm = 'y'
    if not auto_execute:
        logging.info("已强制设置为自动执行模式")
        
        success = 0  # 添加这一行
        failed = 0   # 添加这一行
        
        for op in rename_ops:
            try:
                if create_backup_flag and backup_dir:
                    if not create_backup(op['old_path'], backup_dir):
                        logging.warning(f"备份失败，但继续重命名操作...")
                
                os.rename(op['old_path'], op['new_path'])
                
                rel_path = os.path.relpath(op['directory'], base_dir)
                logging.info(f'[成功] "{op["old_name"]}" -> "{op["new_name"]}"')
                logging.info(f'       路径: {rel_path}')
                success += 1
                
            except Exception as e:
                logging.error(f'[失败] "{op["old_name"]}" -> "{op["new_name"]}"，错误：{e}')
                failed += 1
        
        logging.info(f"\n重命名完成：成功 {success} 个，失败 {failed} 个。")
        logging.info(f"\n统计信息：")
        logging.info(f"  总映射规则: {len(raw_mappings)}")
        logging.info(f"  成功重命名: {success}")
        logging.info(f"  文件不存在: {len(not_found)}")
        logging.info(f"  文件名冲突: {len(conflicts)}")
        logging.info(f"  目录无效: {len(invalid_paths)}")
        if found_by_recursive:
            logging.info(f"  递归找到: {len(found_by_recursive)}")
    else:
        logging.info("已取消操作。")
    
    # 程序结束前暂停
    input("\n按回车键退出...")

if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n程序被用户中断")
        input("\n按回车键退出...")
    except Exception as e:
        print(f"\n程序运行出错: {e}")
        import traceback
        traceback.print_exc()
        input("\n按回车键退出...")